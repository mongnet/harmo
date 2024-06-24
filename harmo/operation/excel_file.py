#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Created by hu on 2018/4/29

from openpyxl import load_workbook
from pathlib2 import Path
from harmo.base_utils import file_is_exist

class OperationExcel:
    '''
    操作excel
    '''
    def __init__(self,file_path,sheetID=0):
        '''
        :param file_path: excel文件路径
        :param sheetId: sheet ID基于索引获取sheet对象
        '''
        if Path(file_path).suffix in [".xlsx",".xls",".xlsm",".xltm"]:
            self.file = file_is_exist(file_path)
        else:
            raise RuntimeError("The file format must be excel")
        self.sheet = self._get_data(sheetID)

    def _get_data(self,sheetID):
        '''
        打开excel,获取指定sheet的数据
        :return: 返回sheet数据
        '''
        wd = load_workbook(self.file)
        return wd.worksheets[sheetID]

    def get_rows(self):
        '''
        返回总行数
        :return:
        '''
        return self.sheet.rows

    def get_columns(self):
        '''
        返回总列数
        :return:
        '''
        return self.sheet.columns

    def get_cell(self,*args):
        '''
        返回指定列
        :return:
        '''
        if len(args) == 2:
            if isinstance(args[0],int) and isinstance(args[1],int):
                return self.sheet.cell(args[0],args[1]).value
            else:
                raise ValueError("参数必需是int类型")
        elif len(args) == 1:
            if isinstance(args[0],str):
                return self.sheet[args[0]].value
            else:
                raise ValueError("参数必需是字符串类型")
        else:
            raise ValueError("获取指定列数据只支持 1,1 或 A1 方式传参")

if __name__ == '__main__':
    oper = OperationExcel(file_path="../template/data/WBS.xlsx",sheetID=0)
    # print(oper.get_rows())
    # print(oper.get_columns())
    # print(oper.get_cell(1,1))
    print(oper.get_cell(2,1))
    print(oper.get_cell(2,2))
    print(oper.get_cell("A1"))
